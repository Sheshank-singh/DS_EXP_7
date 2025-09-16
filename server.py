# server_lb.py – Main server with capacity limit and backup offload
import time, datetime, threading
from pathlib import Path
from typing import Dict, Set
from xmlrpc.server import SimpleXMLRPCServer
from socketserver import ThreadingMixIn
import xmlrpc.client, http.client
from openpyxl import Workbook, load_workbook

SERVER_HOST, SERVER_PORT = "0.0.0.0", 9000
TEACHER_HOST, TEACHER_PORT = "127.0.0.1", 9001
CLIENT_HOST, CLIENT_PORT = "127.0.0.1", 9002
BACKUP_HOST, BACKUP_PORT = "127.0.0.1", 9010  # backup server

PROCESSING_CAPACITY = 3  # main server can do 3 concurrent MCQ finalisations

# --- RPC proxy helpers ---
class TimeoutTransport(xmlrpc.client.Transport):
    def __init__(self, timeout=5.0):
        super().__init__(); self._timeout=timeout
    def make_connection(self, host):
        return http.client.HTTPConnection(host, timeout=self._timeout)

def proxy(url): 
    return xmlrpc.client.ServerProxy(url, allow_none=True, transport=TimeoutTransport())

teacher_proxy = proxy(f"http://{TEACHER_HOST}:{TEACHER_PORT}/")
client_proxy  = proxy(f"http://{CLIENT_HOST}:{CLIENT_PORT}/")
backup_proxy  = proxy(f"http://{BACKUP_HOST}:{BACKUP_PORT}/")

class ThreadingXMLRPCServer(ThreadingMixIn, SimpleXMLRPCServer):
    daemon_threads=True

# ---- state ----
students_registry: Dict[str,str] = {}  # roll -> student xmlrpc URL
student_flags: Dict[str,int]={}
terminated_students:Set[str]=set()
roll_to_name={"1":"Swaroop","2":"Tanisha","3":"Siddhesh","4":"Ayush","5":"Nidhi"}

local_time=None

# MCQ data
MCQ_QUESTIONS={
    1:{"q":"Which protocol is used for time synchronization?","options":["Lamport","Berkeley","Ricart-Agrawala","HTTP"],"answer":2},
    2:{"q":"Which algorithm ensures mutual exclusion for ISA marks?","options":["Token Ring","Ricart-Agrawala","Paxos","Bully"],"answer":2},
    3:{"q":"Which library writes Excel files?","options":["pandas","openpyxl","xlrd","xlsxwriter"],"answer":2},
    4:{"q":"What is the exam duration (seconds)?","options":["20","60","300","600"],"answer":3},
    5:{"q":"Which RPC mechanism is used between nodes?","options":["gRPC","XML-RPC","REST","WebSocket"],"answer":2},
    6:{"q":"A warning (first cheating) reduces MCQ marks to what percent?","options":["100%","80%","50%","0%"],"answer":2},
    7:{"q":"If a student receives 2 warnings, MCQ marks become:","options":["100%","80%","50%","0%"],"answer":4},
    8:{"q":"Total MCQ marks possible:","options":["50","70","100","120"],"answer":3},
    9:{"q":"Who coordinates registration of student peer URLs?","options":["Teacher","Client","Server","Student"],"answer":3},
    10:{"q":"Which data structure logs RA intents?","options":["list","heap","set","dict"],"answer":2}
}

mcq_lock=threading.Lock()
mcq_active=False  # exam running?
mcq_student_answers:Dict[str,Dict[int,int]]={}
mcq_submitted_students:Set[str]=set()
mcq_final_scores:Dict[str,int]={}

processing_semaphore=threading.BoundedSemaphore(PROCESSING_CAPACITY)
processing_now:Set[str]=set()
forwarded_pending:Set[str]=set()
processing_lock=threading.Lock()

excel_path=Path("results.xlsx")

# helper to convert int keys to str for XML-RPC
def _stringify_keys(d: dict) -> dict:
    return {str(k): v for k, v in d.items()}

# ---- functions ----
def register_student(roll, student_url):
    students_registry[str(roll)] = student_url
    print(f"[Server] Registered student {roll} at {student_url}")
    return True

def input_time():
    global local_time
    s=input("[Server] Enter current time (HH-MM-SS): ")
    local_time=datetime.datetime.strptime(s,"%H-%M-%S")
    print("[Server] Local time set:",local_time.strftime("%H-%M-%S"))
    start_mcq()
    return True

def get_time(): 
    return local_time.strftime('%H-%M-%S') if local_time else ""

def start_synchronization():
    print("\n[Server] Starting time synchronization ...\n"+"-"*60)
    try:
        teacher_proxy.send_time()
    except Exception as e:
        print("[Server] Could not sync teacher:", e)
    try:
        client_proxy.send_time()
    except Exception as e:
        print("[Server] Could not sync client:", e)
    return True

def start_mcq():
    global mcq_active
    with mcq_lock:
        mcq_active=True
    print("[Server] MCQ exam started; notifying students...")
    for roll, url in students_registry.items():
        try:
            xmlrpc.client.ServerProxy(url, allow_none=True).start_mcq()
        except Exception as e:
            print(f"[Server] Could not notify student {roll}: {e}")
    threading.Timer(30.0, exam_completed).start()
    return True

def get_mcq_active():
    with mcq_lock: return mcq_active

def get_question_for_student(roll,qnum:int):
    q=MCQ_QUESTIONS.get(int(qnum))
    if not q: return {}
    return {"qnum":int(qnum),"q":q["q"],"options":q["options"]}

def submit_mcq_answer(roll,qnum,ans):
    with mcq_lock:
        mcq_student_answers.setdefault(str(roll),{})[int(qnum)]=int(ans)
    print(f"[Server] recorded ans roll={roll} q={qnum} ans={ans}")
    return True

def exam_completed():
    global mcq_active
    print("[Server] Exam duration over – auto-submitting MCQs...")
    with mcq_lock:
        mcq_active = False
    for roll in list(students_registry.keys()):
        try:
            submit_mcq_final(roll)
        except Exception as e:
            print(f"[Server] Could not auto-submit roll {roll}: {e}")
    print("[Server] Broadcasting to students to start ISA marks entry...")
    for roll, url in students_registry.items():
        try:
            xmlrpc.client.ServerProxy(url, allow_none=True).ask_to_request()
        except Exception as e:
            print(f"[Server] Could not notify student {roll} to start ISA: {e}")
    return True

def _compute_score(answers,flags):
    raw=0
    for qnum,qdef in MCQ_QUESTIONS.items():
        if int(answers.get(qnum,0)or 0)==qdef["answer"]: raw+=10
    if flags>=2:final=0
    elif flags==1:final=int(raw*0.8)
    else:final=raw
    return raw,final

def _finalize_local(roll):
    try:
        print(f"[Server] Processing LOCALLY roll {roll}")
        ans=mcq_student_answers.get(roll,{})
        flags=student_flags.get(roll,0)
        raw,final=_compute_score(ans,flags)
        time.sleep(1.0)
        with mcq_lock:
            mcq_final_scores[roll]=final;mcq_submitted_students.add(roll)
        print(f"[Server] Local done roll={roll} raw={raw} final={final}")
        teacher_proxy.update_mcq_marks(str(roll),int(final))
        wb=load_workbook(excel_path) if excel_path.exists() else Workbook()
        ws=wb.active
        found=False
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value)==roll:
                ws.cell(row=row[0].row,column=4,value=int(final))
                found=True;break
        if not found:
            ws.append([roll,roll_to_name.get(roll,f"Student{roll}"),"NA",int(final),"NA"])
        wb.save(excel_path)
    finally:
        with processing_lock:processing_now.discard(roll)
        processing_semaphore.release()

def submit_mcq_final(roll):
    roll=str(roll)
    with mcq_lock:
        if roll in mcq_submitted_students: return True
    if processing_semaphore.acquire(blocking=False):
        with processing_lock:processing_now.add(roll)
        threading.Thread(target=_finalize_local,args=(roll,),daemon=True).start()
        print(f"[Server] Accepted roll {roll} local (cap {len(processing_now)}/3)")
        return True
    else:
        ans=mcq_student_answers.get(roll,{})
        flags=student_flags.get(roll,0)
        print(f"[Server] Capacity full -> forward roll {roll}")
        try:
            backup_proxy.process_forwarded_submission(roll,_stringify_keys(ans),int(flags))
            with processing_lock:forwarded_pending.add(roll)
        except Exception as e:
            print(f"[Server] Could not auto-submit roll {roll}: {e}")
        return True

def backup_result(roll,final_score):
    roll=str(roll)
    print(f"[Server] got BACKUP result roll {roll}={final_score}")
    with mcq_lock:
        mcq_final_scores[roll]=int(final_score);mcq_submitted_students.add(roll)
    teacher_proxy.update_mcq_marks(str(roll),int(final_score))
    with processing_lock:forwarded_pending.discard(roll)
    return True

def run_server():
    srv=ThreadingXMLRPCServer((SERVER_HOST,SERVER_PORT),allow_none=True,logRequests=False)
    srv.register_function(register_student,"register_student")
    srv.register_function(start_mcq,"start_mcq")
    srv.register_function(input_time,"input_time")
    srv.register_function(get_time,"get_time")
    srv.register_function(start_synchronization,"start_synchronization")
    srv.register_function(get_mcq_active,"get_mcq_active")
    srv.register_function(exam_completed, "exam_completed")
    srv.register_function(get_question_for_student,"get_question_for_student")
    srv.register_function(submit_mcq_answer,"submit_mcq_answer")
    srv.register_function(submit_mcq_final,"submit_mcq_final")
    srv.register_function(backup_result,"backup_result")
    print("[Server] running with load-balancing on port 9000 ...")
    srv.serve_forever()

if __name__=="__main__":
    run_server()
