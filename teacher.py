# teacher.py
from xmlrpc.server import SimpleXMLRPCServer
from socketserver import ThreadingMixIn
import datetime
import threading
import xmlrpc.client
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    raise SystemExit("Please install openpyxl: pip install openpyxl")

class ThreadingXMLRPCServer(ThreadingMixIn, SimpleXMLRPCServer):
    daemon_threads = True

# sample student data (preserved)
students = {
    "1": {"name": "Swaroop", "marks": 100, "flag": 0, "mcq": None},
    "2": {"name": "Tanisha", "marks": 100, "flag": 0, "mcq": None},
    "3": {"name": "Siddhesh", "marks": 100, "flag": 0, "mcq": None},
    "4": {"name": "Ayush", "marks": 100, "flag": 0, "mcq": None},
    "5": {"name": "Nidhi", "marks": 100, "flag": 0, "mcq": None},
}

local_time = None
excel_path = Path("results.xlsx")
_write_lock = threading.Lock()
results_ready = False

def input_time():
    global local_time
    s = input("[Teacher] Enter local time (HH-MM-SS): ")
    local_time = datetime.datetime.strptime(s, "%H-%M-%S")
    print("[Teacher] Time set.")
    return True

def calculate_cv(server_time_str):
    global local_time
    server_time = datetime.datetime.strptime(server_time_str, "%H-%M-%S")
    cv = (local_time - server_time).total_seconds()
    proxy = xmlrpc.client.ServerProxy("http://127.0.0.1:9000/", allow_none=True)
    proxy.receive_cv("Teacher", cv)
    return True

def apply_adjustment(adj):
    global local_time
    if local_time is not None:
        local_time = local_time + datetime.timedelta(seconds=float(adj))
        print(f"[Teacher] Adjusted local time: {local_time.strftime('%H-%M-%S')}")
    return True

def start_exam():
    print("[Teacher] Received start_exam()")
    return True

def deduct_marks(roll, flag):
    if roll in students:
        students[roll]["flag"] = flag
        if flag == 1:
            students[roll]["marks"] = int(students[roll]["marks"] * 0.8)
        elif flag == 2:
            students[roll]["marks"] = 0
    return True




def update_mcq_marks(roll, mcq_marks):
    """
    Server calls this after MCQ finalization for each student.
    Store MCQ marks and update Excel immediately for clarity.
    """
    roll = str(roll)
    with _write_lock:
        if roll not in students:
            # add a new entry if teacher didn't have this student
            students[roll] = {
                "name": f"Student{roll}",
                "marks": 0,
                "flag": 0,
                "mcq": int(mcq_marks),
            }
        else:
            students[roll]["mcq"] = int(mcq_marks)
        print(f"[Teacher] Received MCQ marks for roll {roll}: {mcq_marks}")

        # Update Excel: create file & header if missing
        try:
            if not excel_path.exists():
                wb = Workbook()
                ws = wb.active
                ws.append(["Roll", "Name", "Marks", "MCQ", "ISA"])
                for r, info in students.items():
                    ws.append([
                        r,
                        info.get("name", f"Student{r}"),
                        info.get("marks", "NA"),
                        info.get("mcq", "NA"),
                        "NA",
                    ])
                wb.save(excel_path)
            else:
                wb = load_workbook(excel_path)
                ws = wb.active
                updated = False
                for row in ws.iter_rows(min_row=2):
                    if row and str(row[0].value) == str(roll):
                        # ensure columns for MCQ exist
                        while ws.max_column < 4:
                            ws.cell(row=1, column=ws.max_column + 1, value=None)
                        ws.cell(row=row[0].row, column=4, value=int(mcq_marks))
                        updated = True
                        break
                if not updated:
                    ws.append([
                        roll,
                        students[roll].get("name", f"Student{roll}"),
                        students[roll].get("marks", "NA"),
                        int(mcq_marks),
                        "NA",
                    ])
                wb.save(excel_path)
        except Exception as e:
            print("[Teacher] ERROR updating Excel:", e)

    # ✅ mark results as ready
    global results_ready
    results_ready = True
    return True



def get_results():
    # Return tuples: (roll, name, examMarks, mcq) - mcq may be None
    ret = []
    for r, info in students.items():
        ret.append((r, info.get("name"), info.get("marks"), info.get("mcq", "NA")))
    return ret



def run_teacher():
    server = ThreadingXMLRPCServer(("0.0.0.0", 9001), allow_none=True, logRequests=False)
    server.register_function(input_time, "input_time")
    server.register_function(calculate_cv, "calculate_cv")
    server.register_function(apply_adjustment, "apply_adjustment")
    server.register_function(start_exam, "start_exam")
    server.register_function(deduct_marks, "deduct_marks")
    server.register_function(get_results, "get_results")
    server.register_function(release_results, "release_results")

    server.register_function(update_mcq_marks, "update_mcq_marks")
    print("[Teacher] Running on port 9001...")
    server.serve_forever()

def release_results():
    import pandas as pd
    from openpyxl import load_workbook

    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)  # (Roll, Name, Marks, MCQ, ISA)
    except Exception as e:
        print("[Teacher] ERROR reading Excel:", e)
        return False

    print("[Teacher] Results released to students.")

    proxy = xmlrpc.client.ServerProxy("http://127.0.0.1:9000/", allow_none=True)
    proxy.announce_results(data)
    return True


if __name__ == "__main__":
    import threading
    import time

    # Start RPC server in background
    threading.Thread(target=run_teacher, daemon=True).start()

    # Wait until results are ready
    while not results_ready:
        time.sleep(1)

    # Teacher manual release loop
    while True:
        choice = input(
            "[Teacher] Do you want to release results to students? (y/n/exit): "
        ).strip().lower()
        if choice == "y":
            release_results()
        elif choice == "n":
            print("[Teacher] Results not released yet.")
        elif choice == "exit":
            print("[Teacher] Exiting teacher console.")
            break
        else:
            print("[Teacher] Please enter y/n/exit.")
